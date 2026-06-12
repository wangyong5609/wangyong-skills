#!/usr/bin/env python3
"""Download cover images from a source Excel file.

Expected columns: ID, 作者, 标题, 发布时间, 点赞, 分享, 评论, 收藏, 推荐, 封面.
"""

from __future__ import annotations

import argparse
import json
import pathlib
import time
import urllib.request


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download covers from an Excel table.")
    parser.add_argument("xlsx", help="Path to the source .xlsx file")
    parser.add_argument("--out", default="downloads/yingshi-hurricane-covers", help="Output image directory")
    parser.add_argument("--manifest", default="downloads/yingshi-hurricane-covers-manifest.jsonl", help="Manifest JSONL path")
    parser.add_argument("--sleep", type=float, default=0.08, help="Delay between downloads")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency: openpyxl. Use the bundled Codex Python runtime or install openpyxl.") from exc

    xlsx = pathlib.Path(args.xlsx)
    out = pathlib.Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    manifest = pathlib.Path(args.manifest)
    manifest.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    required = ["ID", "作者", "标题", "发布时间", "点赞", "分享", "评论", "收藏", "推荐", "封面"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Missing columns: {', '.join(missing)}")

    request_headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.douyin.com/"}
    rows: list[dict[str, object]] = []
    ok = 0
    failed = 0

    for row_no in range(2, ws.max_row + 1):
        video_id = str(ws.cell(row_no, idx["ID"]).value)
        url = ws.cell(row_no, idx["封面"]).value
        if not url:
            continue
        image_path = out / f"{row_no - 1:03d}_{video_id}.jpg"
        record: dict[str, object] = {
            "row": row_no,
            "index": row_no - 1,
            "id": video_id,
            "author": ws.cell(row_no, idx["作者"]).value,
            "title": ws.cell(row_no, idx["标题"]).value,
            "published_at": ws.cell(row_no, idx["发布时间"]).value,
            "likes": ws.cell(row_no, idx["点赞"]).value,
            "shares": ws.cell(row_no, idx["分享"]).value,
            "comments": ws.cell(row_no, idx["评论"]).value,
            "favorites": ws.cell(row_no, idx["收藏"]).value,
            "recommended": ws.cell(row_no, idx["推荐"]).value,
            "cover_url": url,
            "file": str(image_path),
        }

        if image_path.exists() and image_path.stat().st_size > 1000:
            record["download_status"] = "exists"
            record["bytes"] = image_path.stat().st_size
            ok += 1
        else:
            try:
                req = urllib.request.Request(str(url), headers=request_headers)
                with urllib.request.urlopen(req, timeout=20) as resp:
                    data = resp.read()
                image_path.write_bytes(data)
                record["download_status"] = "ok"
                record["bytes"] = len(data)
                ok += 1
                time.sleep(args.sleep)
            except Exception as exc:  # noqa: BLE001 - record per-row failures for manifest.
                record["download_status"] = "failed"
                record["error"] = repr(exc)
                failed += 1
        rows.append(record)

    with manifest.open("w", encoding="utf-8") as fh:
        for record in rows:
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")

    print(json.dumps({"rows": len(rows), "downloaded_or_existing": ok, "failed": failed, "out": str(out), "manifest": str(manifest)}, ensure_ascii=False, indent=2))
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
